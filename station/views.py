from django.shortcuts import render
from django.views import View
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import redirect
from django.utils import timezone
from django.db.models import Q
from django.http import HttpResponseForbidden
from inventory_app.permissions import IsManager
from station.forms import FleetOrderForm, StationInspectionForm, VehicleInspectionForm
from station.models import Fleet_order, Station, Station_inspection, Vehicle_inspection
from django.contrib import messages
import json
from django.http import JsonResponse
import openpyxl
from django.utils.html import escape
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin


class station_view(LoginRequiredMixin, TemplateView):
    template_name = "station_base.html"
    login_url = (
        "authentication:login"  # Redirect to your login page if not authenticated
    )


class report_view(View):

    permission_classes = [IsAuthenticated, IsManager]

    def dispatch(self, request, *args, **kwargs):
        # Manually check permissions before executing the view logic
        for permission in self.permission_classes:
            permission_instance = permission()
            if not permission_instance.has_permission(request, self):
                return redirect("authentication:login")
        return super().dispatch(request, *args, **kwargs)

    template_name = "report.html"

    def get(self, request, station_number):
        active_tab = request.GET.get("tab", "order_report")  # Default to 'order_report'

        # Retrieve start and end dates from the GET request
        start_date_str = request.GET.get("start_date", None)
        end_date_str = request.GET.get("end_date", None)

        # Handle date parsing
        if start_date_str and end_date_str:
            try:
                start_date = timezone.datetime.strptime(
                    start_date_str, "%Y-%m-%d"
                ).date()
                end_date = timezone.datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                start_date = end_date = (
                    timezone.now().date()
                )  # Default to today if parsing fails
        else:
            start_date = end_date = (
                timezone.now().date()
            )  # Default to today if no date provided

        # Initialize context
        context = {
            "station_number": station_number,
            "active_tab": active_tab,
            "start_date": start_date,
            "end_date": end_date,
        }

        # Populate data for the "order_report" tab
        if active_tab == "order_report":
            fleet_orders = Fleet_order.objects.filter(
                station=station_number, type="fleet", date__range=(start_date, end_date)
            )
            office_orders = Fleet_order.objects.filter(
                station=station_number,
                type="office supplies",
                date__range=(start_date, end_date),
            )

            context["fleet_orders"] = fleet_orders
            context["fleet_total_orders"] = fleet_orders.count()
            context["fleet_pending_orders"] = fleet_orders.filter(
                status="pending approval"
            ).count()
            context["fleet_approved_orders"] = fleet_orders.filter(
                status="approved"
            ).count()
            context["fleet_rejected_orders"] = fleet_orders.filter(
                status="rejected"
            ).count()

            context["office_orders"] = office_orders
            context["office_total_orders"] = office_orders.count()
            context["office_pending_orders"] = office_orders.filter(
                status="pending approval"
            ).count()
            context["office_approved_orders"] = office_orders.filter(
                status="approved"
            ).count()
            context["office_rejected_orders"] = office_orders.filter(
                status="rejected"
            ).count()

        elif active_tab == "vehicle_report":
            truck_inspection = Vehicle_inspection.objects.filter(
                station=station_number,
                vehicle__vehicle_type="truck",
                date__range=(start_date, end_date),
            )
            trailer_inspection = Vehicle_inspection.objects.filter(
                station=station_number,
                vehicle__vehicle_type="trailer",
                date__range=(start_date, end_date),
            )

            context["truck_inspection"] = truck_inspection
            context["truck_total_orders"] = truck_inspection.count()
            context["truck_maintenance"] = truck_inspection.filter(
                type="regular maintenance"
            ).count()
            context["truck_repairs"] = truck_inspection.filter(type="repair").count()
            context["truck_inspections"] = truck_inspection.filter(
                type="inspection"
            ).count()

            context["trailer_inspection"] = trailer_inspection
            context["trailer_total_orders"] = trailer_inspection.count()
            context["trailer_maintenance"] = trailer_inspection.filter(
                type="regular maintenance"
            ).count()
            context["trailer_repairs"] = trailer_inspection.filter(
                type="repair"
            ).count()
            context["trailer_inspections"] = trailer_inspection.filter(
                type="inspection"
            ).count()

        elif active_tab == "station_overview":
            context["station1_overview"] = Station_inspection.objects.filter(
                Q(inventory_status="Major inventory issues")
                | Q(inventory_status="Some tools missing"),
                station=1,
                date__range=(start_date, end_date),
            ).last()
            context["station2_overview"] = Station_inspection.objects.filter(
                Q(inventory_status="Major inventory issues")
                | Q(inventory_status="Some tools missing"),
                station=2,
                date__range=(start_date, end_date),
            ).last()

        elif active_tab == "station_summary":
            context["station_details"] = Station_inspection.objects.filter(
                station=station_number, date__range=(start_date, end_date)
            )

        return render(request, self.template_name, context)


class station_inspection_view(View):

    permission_classes = [IsAuthenticated, IsManager]

    def dispatch(self, request, *args, **kwargs):
        # Manually check permissions before executing the view logic
        for permission in self.permission_classes:
            permission_instance = permission()
            if not permission_instance.has_permission(request, self):
                return redirect("authentication:login")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, station_number):
        form = StationInspectionForm()
        return render(
            request,
            "station_inspection.html",
            {"form": form, "station_number": station_number},
        )

    def post(self, request, station_number):

        form = StationInspectionForm(request.POST)

        if form.is_valid():
            inspection = form.save(commit=False)

            # Handle the inspection date (default to today if not provided)
            selected_date_str = request.POST.get("date")
            if selected_date_str:
                selected_date = timezone.datetime.strptime(
                    selected_date_str, "%Y-%m-%d"
                ).date()
            else:
                selected_date = timezone.now().date()

            station = Station.objects.get(id=station_number)

            # Assign the selected date to the inspection
            inspection.date = selected_date
            inspection.station = station
            inspection.saved_on = timezone.now()
            inspection.submitted_by = request.user.username
            inspection.save()

            messages.success(request, "Station Inspection Completed Successfully!")
            return redirect("station_inspection", station_number=station_number)
        else:
            messages.error(request, form.errors)
        return render(
            request,
            "station_inspection.html",
            {"form": form, "station_number": station_number},
        )


class vehicle_inspection_view(View):

    permission_classes = [IsAuthenticated, IsManager]

    def dispatch(self, request, *args, **kwargs):
        # Manually check permissions before executing the view logic
        for permission in self.permission_classes:
            permission_instance = permission()
            if not permission_instance.has_permission(request, self):
                return redirect("authentication:login")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, station_number, vehicle):
        form = VehicleInspectionForm(vehicle_type=vehicle)
        Vehicle_inspections = Vehicle_inspection.objects.filter(
            station=station_number, vehicle__vehicle_type=vehicle
        )
        return render(
            request,
            "vehicle_inspection.html",
            {
                "form": form,
                "station_number": station_number,
                "vehicle": vehicle,
                "Vehicle_inspections": Vehicle_inspections,
            },
        )

    def post(self, request, station_number, vehicle):

        form = VehicleInspectionForm(request.POST)

        if form.is_valid():
            inspection = form.save(commit=False)

            station = Station.objects.get(id=station_number)

            inspection.station = station
            inspection.saved_on = timezone.now()
            inspection.submitted_by = request.user.username
            inspection.save()

            messages.success(request, f"{vehicle} Inspection Completed Successfully!")
            return redirect(
                "vehicle_inspection", station_number=station_number, vehicle=vehicle
            )
        else:
            messages.error(request, form.errors)

        return render(
            request,
            "vehicle_inspection.html",
            {"form": form, "station_number": station_number, "vehicle": vehicle},
        )


class order_view(View):

    permission_classes = [IsAuthenticated, IsManager]

    def dispatch(self, request, *args, **kwargs):
        # Manually check permissions before executing the view logic
        for permission in self.permission_classes:
            permission_instance = permission()
            if not permission_instance.has_permission(request, self):
                return redirect("authentication:login")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, station_number, type):
        form = FleetOrderForm()
        fleet_orders = Fleet_order.objects.filter(type=type, station__id=station_number)
        return render(
            request,
            "order.html",
            {
                "form": form,
                "station_number": station_number,
                "type": type,
                "fleet_orders": fleet_orders,
            },
        )

    def post(self, request, station_number, type):

        form = FleetOrderForm(request.POST)

        if form.is_valid():
            order = form.save(commit=False)

            station = Station.objects.get(id=station_number)

            order.station = station
            order.saved_on = timezone.now()
            order.submitted_by = request.user.username
            order.type = type
            order.save()

            messages.success(request, f"{type} Order Submitted Successfully!")
            return redirect("order", station_number=station_number, type=type)
        else:
            messages.error(request, form.errors)

        return render(
            request,
            "order.html",
            {"form": form, "station_number": station_number, "type": type},
        )


def get_merged_cell_ranges(sheet):
    """Get merged cell range details as a dictionary with coordinates."""
    merged_cells = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        merged_cells[(min_row, min_col)] = {
            "rowspan": max_row - min_row + 1,
            "colspan": max_col - min_col + 1,
        }
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) != (min_row, min_col):
                    merged_cells[(row, col)] = None
    return merged_cells


def get_hex_color(cell, is_background=True):
    """Convert OpenPyXL color to hex format, handling indexed and theme colors correctly.
    By default, this checks the background color (is_background=True),
    but you can use is_background=False to get text color.
    """
    color = cell.fill.fgColor if is_background else cell.font.color
    if not color:
        return "#000000"  # Default to black if no color is found.

    if color.type == "rgb":
        return f"#{color.rgb[-6:]}"

    if color.type == "theme":
        indexed_colors = {
            8: "#000000",
            9: "#FFFFFF",
            10: "#FF0000",
            11: "#00FF00",
            12: "#0000FF",
            13: "#FFFF00",
            14: "#4F81BD",
            15: "B71C1C",
            16: "C0504D",
            17: "c0504d",
        }
        return indexed_colors.get(color.indexed, "#4f81bd")

    return "#000000"  # Default to black if type is not matched


def excel_view(request, station_number):
    EXCEL_FILE_PATH = ""

    if station_number == 1:
        EXCEL_FILE_PATH = "station/station_1_.xlsx"
    else:
        EXCEL_FILE_PATH = "station/station_2_.xlsx"

    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
    sheet = wb.active
    merged_cells = get_merged_cell_ranges(sheet)

    # Find the last row and column with data
    last_row = sheet.max_row
    last_col = sheet.max_column

    # Strip out any empty rows or columns at the bottom
    while last_row > 0 and all(
        sheet.cell(row=last_row, column=col).value is None
        for col in range(1, last_col + 1)
    ):
        last_row -= 1

    while last_col > 0 and all(
        sheet.cell(row=row, column=last_col).value is None
        for row in range(1, last_row + 1)
    ):
        last_col -= 1

    # Build the HTML table for only the used rows and columns
    html_table = '<table class="excel-table">'
    for row_idx in range(1, last_row + 1):  # Only iterate up to last_row
        html_table += "<tr>"
        for col_idx in range(1, last_col + 1):  # Only iterate up to last_col
            cell = sheet.cell(row=row_idx, column=col_idx)
            if (row_idx, col_idx) in merged_cells and merged_cells[
                (row_idx, col_idx)
            ] is None:
                continue

            value = escape(str(cell.value)) if cell.value else ""
            bg_color = get_hex_color(cell, is_background=True)  # Background color
            text_color = get_hex_color(cell, is_background=False)  # Text color
            rowspan = merged_cells.get((row_idx, col_idx), {}).get("rowspan", 1)
            colspan = merged_cells.get((row_idx, col_idx), {}).get("colspan", 1)

            html_table += (
                f'<td contenteditable="true" data-row="{row_idx-1}" data-col="{col_idx-1}" '
                f'style="background-color: {bg_color};"'
            )

            if rowspan > 1:
                html_table += f' rowspan="{rowspan}"'
            if colspan > 1:
                html_table += f' colspan="{colspan}"'

            html_table += f">{value}</td>"

        html_table += "</tr>"
    html_table += "</table>"

    return render(
        request,
        "station_layout.html",
        {"excel_html": html_table, "station_number": station_number},
    )


def save_excel_changes(request, station_number):

    EXCEL_FILE_PATH = ""

    if station_number == 1:
        EXCEL_FILE_PATH = "station/station_1_.xlsx"
    else:
        EXCEL_FILE_PATH = "station/station_2_.xlsx"

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            updates = data.get("data", [])

            if not updates:
                return JsonResponse({"error": "No data received"}, status=400)

            wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = wb.active

            for change in updates:
                row = int(change["row"]) + 1
                col = int(change["col"]) + 1
                sheet.cell(row=row, column=col, value=change["value"])

            wb.save(EXCEL_FILE_PATH)

            return JsonResponse({"message": "Changes saved successfully!"})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request"}, status=400)
